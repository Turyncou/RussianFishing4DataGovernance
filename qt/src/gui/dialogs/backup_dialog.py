"""Backup and restore dialog with Excel export support"""
import os
import shutil
import json
from datetime import datetime
from typing import List
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QScrollArea, QWidget, QButtonGroup, QRadioButton, QMessageBox,
    QFileDialog, QFrame
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont

from src.data.persistence import list_backups, create_auto_backup
from src.core.models import (
    ActivityCharacter, BaitConsumption, StorageCharacter, AccountCredential,
    ActivityType
)


class BackupRestoreDialog(QDialog):
    """Dialog for backup and restore"""

    def __init__(self, parent, data_dir: str, backup_dir: str):
        super().__init__(parent)
        self.data_dir = data_dir
        self.backup_dir = backup_dir
        self.setWindowTitle("备份与恢复")
        self.setFixedSize(550, 450)
        self.setModal(True)

        self._create_widgets()
        self._load_backups()

    def _create_widgets(self):
        """Create the dialog widgets"""
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)

        # Title
        title = QLabel("💾 备份与恢复 / 数据导出")
        title.setFont(QFont("Segoe UI", 18, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Auto backup section
        auto_frame = QWidget()
        auto_frame.setStyleSheet("QWidget { background-color: #252525; border-radius: 12px; }")
        auto_layout = QVBoxLayout(auto_frame)
        auto_layout.setContentsMargins(10, 10, 10, 10)

        section_label = QLabel("自动备份")
        section_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        auto_layout.addWidget(section_label)

        # Scroll area for backups
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)
        self.backup_container = QWidget()
        self.backup_layout = QVBoxLayout(self.backup_container)
        self.backup_layout.setSpacing(3)
        scroll_area.setWidget(self.backup_container)
        scroll_area.setFixedHeight(150)
        auto_layout.addWidget(scroll_area)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(5)

        create_btn = QPushButton("➕ 创建手动备份")
        create_btn.setFixedWidth(140)
        create_btn.clicked.connect(self.create_manual_backup)
        btn_layout.addWidget(create_btn)

        restore_btn = QPushButton("♻️ 恢复选中备份")
        restore_btn.setFixedWidth(140)
        restore_btn.setStyleSheet("""
            QPushButton {
                background-color: #cc8800;
            }
            QPushButton:hover {
                background-color: #aa6600;
            }
        """)
        restore_btn.clicked.connect(self.restore_selected)
        btn_layout.addWidget(restore_btn)

        delete_btn = QPushButton("🗑️ 删除选中备份")
        delete_btn.setFixedWidth(140)
        delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #cc3333;
            }
            QPushButton:hover {
                background-color: #aa2222;
            }
        """)
        delete_btn.clicked.connect(self.delete_selected)
        btn_layout.addWidget(delete_btn)

        btn_layout.addStretch()
        auto_layout.addLayout(btn_layout)
        layout.addWidget(auto_frame)

        # Export section
        export_frame = QWidget()
        export_frame.setStyleSheet("QWidget { background-color: #252525; border-radius: 12px; }")
        export_layout = QVBoxLayout(export_frame)
        export_layout.setContentsMargins(10, 10, 10, 10)

        export_label = QLabel("导出数据")
        export_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        export_layout.addWidget(export_label)

        export_btn = QPushButton("📊 导出全部数据到 Excel")
        export_btn.setFixedWidth(220)
        export_btn.clicked.connect(self.export_to_excel)
        export_layout.addWidget(export_btn)
        export_layout.addStretch()
        layout.addWidget(export_frame)

        # Close button
        layout.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.setFixedWidth(100)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #888888;
            }
            QPushButton:hover {
                background-color: #666666;
            }
        """)
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignCenter)

        self.button_group = QButtonGroup()
        self.selected_backup = None

    def _clear_backup_list(self):
        """Clear all radio buttons from backup list"""
        for button in self.button_group.buttons():
            self.button_group.removeButton(button)
            button.deleteLater()

    def _load_backups(self):
        """Load backup list"""
        self._clear_backup_list()
        self.selected_backup = None

        backups = list_backups(self.backup_dir)
        if not backups:
            label = QLabel("暂无备份")
            label.setAlignment(Qt.AlignCenter)
            self.backup_layout.addWidget(label)
            return

        for backup in backups:
            # Parse timestamp: backup_20260324_123456 → 2026-03-24 12:34:56
            try:
                parts = backup.split('_')
                date_part = parts[1]
                time_part = parts[2]
                formatted = f"{date_part[:4]}-{date_part[4:6]}-{date_part[6:8]} {time_part[:2]}:{time_part[2:4]}:{time_part[4:6]}"
            except:
                formatted = backup

            rb = QRadioButton(formatted)
            rb.setObjectName(backup)
            rb.toggled.connect(self._on_backup_selected)
            self.button_group.addButton(rb)
            self.backup_layout.addWidget(rb)

        self.backup_layout.addStretch()

    def _on_backup_selected(self, checked):
        """Handle radio button selection"""
        if checked:
            rb = self.sender()
            self.selected_backup = rb.objectName()

    def create_manual_backup(self):
        """Create manual backup"""
        backup_path = create_auto_backup(self.data_dir, self.backup_dir)
        QMessageBox.information(self, "成功", f"备份已创建:\n{backup_path}")
        self._load_backups()

    def restore_selected(self):
        """Restore selected backup"""
        if not self.selected_backup:
            QMessageBox.warning(self, "提示", "请先选择一个备份")
            return

        backup_path = os.path.join(self.backup_dir, self.selected_backup)

        confirm = QMessageBox.question(
            self, "确认恢复",
            f"确定要恢复备份 {self.selected_backup} 吗？\n当前所有数据将会被覆盖！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        # Restore all JSON files from backup
        try:
            for filename in os.listdir(backup_path):
                if filename.endswith('.json') or filename.endswith('.json.key'):
                    src = os.path.join(backup_path, filename)
                    dst = os.path.join(self.data_dir, filename)
                    shutil.copy2(src, dst)
            QMessageBox.information(self, "成功", "恢复成功！\n请重启程序以应用更改")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"恢复失败: {str(e)}")

    def delete_selected(self):
        """Delete selected backup"""
        if not self.selected_backup:
            QMessageBox.warning(self, "提示", "请先选择一个备份")
            return

        backup_path = os.path.join(self.backup_dir, self.selected_backup)

        confirm = QMessageBox.question(
            self, "确认删除",
            f"确定要删除备份 {self.selected_backup} 吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        try:
            shutil.rmtree(backup_path, ignore_errors=True)
            QMessageBox.information(self, "成功", "备份已删除")
            self._load_backups()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"删除失败: {str(e)}")

    def export_to_excel(self):
        """Export all data to Excel file"""
        # Ask for save location
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"rf4_data_export_{timestamp}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出Excel",
            default_filename,
            "Excel 文件 (*.xlsx);;所有文件 (*.*)"
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook

            wb = Workbook()

            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

            # Export activity data
            self._export_activity_data(wb)
            # Export bait data
            self._export_bait_data(wb)
            # Export storage data
            self._export_storage_data(wb)
            # Export credentials data (just account names, no passwords)
            self._export_credentials(wb)

            wb.save(file_path)
            QMessageBox.information(self, "成功", f"数据已导出到:\n{file_path}")
        except ImportError:
            QMessageBox.critical(self, "错误", "需要安装 openpyxl 才能导出Excel\n请运行: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")

    def _export_activity_data(self, wb):
        """Export activity statistics to Excel sheet"""
        ws = wb.create_sheet(title="活动统计")
        # Headers
        ws.append(["角色名称", "活动类型", "今日价值", "今日时长", "总计价值", "总计时长", "目标价值", "目标收入"])

        # Load data from file
        activity_file = os.path.join(self.data_dir, 'activity.json')
        if os.path.exists(activity_file):
            with open(activity_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Handle both old (list) and new (dict with characters) formats
                if isinstance(data, dict) and 'characters' in data:
                    characters_data = data['characters']
                else:
                    characters_data = data if isinstance(data, list) else []

                for char_data in characters_data:
                    # We need to reconstruct the ActivityCharacter manually
                    # because the JSON format isn't the same as the constructor kwargs
                    name = char_data.get('name', '')

                    # Load records to calculate totals
                    grinding_total = 0
                    grinding_duration = 0
                    grinding_today = 0
                    grinding_duration_today = 0
                    grinding_target = 0
                    grinding_income = 0

                    star_total = 0
                    star_duration = 0
                    star_today = 0
                    star_duration_today = 0
                    star_target = 0
                    star_income = 0

                    # We can't easily do the full calculation without loading everything
                    # So just output what we have directly from JSON
                    grinding_goal_data = char_data.get('grinding_goal')
                    if grinding_goal_data:
                        grinding_target = grinding_goal_data.get('target_value', 0)
                        grinding_income = grinding_goal_data.get('total_income', 0)

                    star_goal_data = char_data.get('star_waiting_goal')
                    if star_goal_data:
                        star_target = star_goal_data.get('target_value', 0)
                        star_income = star_goal_data.get('total_income', 0)

                    from datetime import date
                    today = date.today()
                    for record_data in char_data.get('records', []):
                        record_date = date.fromisoformat(record_data.get('date'))
                        is_today = record_date == today
                        if record_data.get('activity_type') == 'grinding':
                            val = record_data.get('silver_count', 0)
                            dur = record_data.get('duration_minutes', 0)
                            grinding_total += val
                            grinding_duration += dur
                            if is_today:
                                grinding_today += val
                                grinding_duration_today += dur
                        else:
                            val = record_data.get('success_count', 0)
                            dur = record_data.get('duration_minutes', 0)
                            star_total += val
                            star_duration += dur
                            if is_today:
                                star_today += val
                                star_duration_today += dur

                    ws.append([
                        name, "搬砖",
                        grinding_today, grinding_duration_today,
                        grinding_total, grinding_duration,
                        grinding_target, grinding_income
                    ])

                    ws.append([
                        name, "蹲星",
                        star_today, star_duration_today,
                        star_total, star_duration,
                        star_target, star_income
                    ])

    def _export_bait_data(self, wb):
        """Export bait data to Excel"""
        ws = wb.create_sheet(title="饵料库存")
        ws.append(["名称", "已购买", "已使用", "剩余"])

        bait_file = os.path.join(self.data_dir, 'bait.json')
        if os.path.exists(bait_file):
            with open(bait_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for item in data:
                    name = item.get('name', '')
                    bought = item.get('total_bought', 0)
                    used = item.get('total_used', 0)
                    remaining = bought - used
                    if remaining < 0:
                        remaining = 0
                    ws.append([name, bought, used, remaining])

    def _export_storage_data(self, wb):
        """Export storage data to Excel"""
        ws = wb.create_sheet(title="存储时长")
        ws.append(["角色名称", "剩余时长(分钟)"])

        storage_file = os.path.join(self.data_dir, 'storage.json')
        if os.path.exists(storage_file):
            with open(storage_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for item in data:
                    name = item.get('name', '')
                    remaining = item.get('remaining_minutes', 0)
                    ws.append([name, remaining])

    def _export_credentials(self, wb):
        """Export account names to Excel (passwords not exported for security)"""
        ws = wb.create_sheet(title="账号列表")
        ws.append(["账号名称"])

        cred_file = os.path.join(self.data_dir, 'credentials.json')
        if os.path.exists(cred_file):
            with open(cred_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for item in data:
                    name = item.get('account_name', '')
                    ws.append([name])
