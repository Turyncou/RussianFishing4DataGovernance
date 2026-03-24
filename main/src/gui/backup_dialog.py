"""Backup and restore dialog with Excel export support"""
import customtkinter as ctk
from tkinter import filedialog
from CTkMessagebox import CTkMessagebox
import os
import shutil
import json
from openpyxl import Workbook
from datetime import datetime
from typing import List

from data.persistence import list_backups, create_auto_backup
from core.models import (
    ActivityCharacter, BaitConsumption, StorageCharacter, AccountCredential,
    ActivityType
)


class BackupRestoreDialog(ctk.CTkToplevel):
    """Dialog for backup and restore"""

    def __init__(self, parent, data_dir: str, backup_dir: str):
        super().__init__(parent)
        self.data_dir = data_dir
        self.backup_dir = backup_dir
        self.title("备份与恢复")
        self.geometry("550x450")
        self.resizable(False, False)
        self.grab_set()

        self.create_widgets()
        self.load_backups()

    def create_widgets(self):
        """Create the dialog widgets"""
        # Title
        title = ctk.CTkLabel(
            self,
            text="💾 备份与恢复 / 数据导出",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title.pack(pady=(15, 10))

        # Auto backup section
        auto_frame = ctk.CTkFrame(self, fg_color="#252525", corner_radius=12)
        auto_frame.pack(fill="x", padx=15, pady=(0, 10))

        ctk.CTkLabel(
            auto_frame,
            text="自动备份",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(pady=(10, 5), anchor="w", padx=10)

        self.backup_listbox = ctk.CTkScrollableFrame(auto_frame, width=500, height=150)
        self.backup_listbox.pack(padx=10, pady=5, fill="both")

        # Buttons for backup list
        btn_frame = ctk.CTkFrame(auto_frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=(5, 10))

        create_btn = ctk.CTkButton(
            btn_frame,
            text="➕ 创建手动备份",
            command=self.create_manual_backup,
            width=120,
            corner_radius=8
        )
        create_btn.pack(side="left", padx=5)

        restore_btn = ctk.CTkButton(
            btn_frame,
            text="♻️ 恢复选中备份",
            command=self.restore_selected,
            width=120,
            corner_radius=8,
            fg_color="#cc8800",
            hover_color="#aa6600"
        )
        restore_btn.pack(side="left", padx=5)

        delete_btn = ctk.CTkButton(
            btn_frame,
            text="🗑️ 删除选中备份",
            command=self.delete_selected,
            width=120,
            corner_radius=8,
            fg_color="#cc3333",
            hover_color="#aa2222"
        )
        delete_btn.pack(side="left", padx=5)

        # Export section
        export_frame = ctk.CTkFrame(self, fg_color="#252525", corner_radius=12)
        export_frame.pack(fill="x", padx=15, pady=(0, 15))

        ctk.CTkLabel(
            export_frame,
            text="导出数据",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(pady=(10, 5), anchor="w", padx=10)

        export_btn = ctk.CTkButton(
            export_frame,
            text="📊 导出全部数据到 Excel",
            command=self.export_to_excel,
            width=180,
            corner_radius=8
        )
        export_btn.pack(padx=10, pady=10)

        # Close button
        close_btn = ctk.CTkButton(
            self,
            text="关闭",
            command=self.destroy,
            width=100,
            corner_radius=8,
            fg_color="#888888",
            hover_color="#666666"
        )
        close_btn.pack(pady=(0, 15))

    def load_backups(self):
        """Load backup list"""
        # Clear existing
        for widget in self.backup_listbox.winfo_children():
            widget.destroy()

        backups = list_backups(self.backup_dir)
        if not backups:
            label = ctk.CTkLabel(
                self.backup_listbox,
                text="暂无备份",
                font=ctk.CTkFont(size=12)
            )
            label.pack(pady=20)
            return

        self.selected_backup_var = ctk.StringVar(value="")
        for backup in backups:
            # Parse timestamp: backup_20260324_123456 → 2026-03-24 12:34:56
            try:
                date_part = backup.split('_')[1]
                time_part = backup.split('_')[2]
                formatted = f"{date_part[:4]}-{date_part[4:6]}-{date_part[6:8]} {time_part[:2]}:{time_part[2:4]}:{time_part[4:6]}"
            except:
                formatted = backup

            rb = ctk.CTkRadioButton(
                self.backup_listbox,
                text=formatted,
                variable=self.selected_backup_var,
                value=backup
            )
            rb.pack(pady=3, anchor="w")

    def create_manual_backup(self):
        """Create manual backup"""
        backup_path = create_auto_backup(self.data_dir, self.backup_dir)
        CTkMessagebox(title="成功", message=f"备份已创建:\n{backup_path}", icon="info")
        self.load_backups()

    def restore_selected(self):
        """Restore selected backup"""
        selected = getattr(self, 'selected_backup_var', None)
        if not selected or not selected.get():
            CTkMessagebox(title="提示", message="请先选择一个备份", icon="warning")
            return

        backup_name = selected.get()
        backup_path = os.path.join(self.backup_dir, backup_name)

        confirm = CTkMessagebox(
            title="确认恢复",
            message=f"确定要恢复备份 {backup_name} 吗？\n当前所有数据将会被覆盖！",
            icon="warning",
            option_1="取消",
            option_2="恢复"
        )
        if confirm.get() != "恢复":
            return

        # Restore all JSON files from backup
        try:
            for filename in os.listdir(backup_path):
                if filename.endswith('.json') or filename.endswith('.json.key'):
                    src = os.path.join(backup_path, filename)
                    dst = os.path.join(self.data_dir, filename)
                    shutil.copy2(src, dst)
            CTkMessagebox(title="成功", message="恢复成功！\n请重启程序以应用更改", icon="info")
            self.destroy()
        except Exception as e:
            CTkMessagebox(title="错误", message=f"恢复失败: {str(e)}", icon="cancel")

    def delete_selected(self):
        """Delete selected backup"""
        selected = getattr(self, 'selected_backup_var', None)
        if not selected or not selected.get():
            CTkMessagebox(title="提示", message="请先选择一个备份", icon="warning")
            return

        backup_name = selected.get()
        backup_path = os.path.join(self.backup_dir, backup_name)

        confirm = CTkMessagebox(
            title="确认删除",
            message=f"确定要删除备份 {backup_name} 吗？",
            icon="warning",
            option_1="取消",
            option_2="删除"
        )
        if confirm.get() != "删除":
            return

        try:
            shutil.rmtree(backup_path, ignore_errors=True)
            CTkMessagebox(title="成功", message="备份已删除", icon="info")
            self.load_backups()
        except Exception as e:
            CTkMessagebox(title="错误", message=f"删除失败: {str(e)}", icon="cancel")

    def export_to_excel(self):
        """Export all data to Excel file"""
        # Ask for save location
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"rf4_data_export_{timestamp}.xlsx"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=default_filename
        )
        if not file_path:
            return

        try:
            wb = Workbook()

            # Export activity data
            self._export_activity_data(wb)
            # Export bait data
            self._export_bait_data(wb)
            # Export storage data
            self._export_storage_data(wb)
            # Export credentials data (just account names, no passwords)
            self._export_credentials(wb)

            wb.save(file_path)
            CTkMessagebox(title="成功", message=f"数据已导出到:\n{file_path}", icon="info")
        except Exception as e:
            CTkMessagebox(title="错误", message=f"导出失败: {str(e)}", icon="cancel")

    def _export_activity_data(self, wb: Workbook):
        """Export activity statistics to Excel sheet"""
        ws = wb.create_sheet(title="活动统计")
        # Headers
        ws.append(["角色名称", "活动类型", "今日价值", "今日时长", "总计价值", "总计时长", "目标价值", "目标收入"])

        # Load data from file
        activity_file = os.path.join(self.data_dir, 'activity.json')
        if os.path.exists(activity_file):
            with open(activity_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for char_data in data:
                    char = ActivityCharacter(**char_data)
                    # Grinding
                    total_value_g, total_duration_g, _ = char.calculate_totals(ActivityType.GRINDING)
                    today_value_g, today_duration_g = char.calculate_today_totals(ActivityType.GRINDING)
                    target_value_g = char.grinding_goal.target_value if char.grinding_goal else 0
                    income_g = char.grinding_goal.total_income if char.grinding_goal else 0
                    ws.append([
                        char.name, "搬砖",
                        today_value_g, today_duration_g,
                        total_value_g, total_duration_g,
                        target_value_g, income_g
                    ])
                    # Star waiting
                    total_value_s, total_duration_s, _ = char.calculate_totals(ActivityType.STAR_WAITING)
                    today_value_s, today_duration_s = char.calculate_today_totals(ActivityType.STAR_WAITING)
                    target_value_s = char.star_waiting_goal.target_value if char.star_waiting_goal else 0
                    income_s = char.star_waiting_goal.total_income if char.star_waiting_goal else 0
                    ws.append([
                        char.name, "蹲星",
                        today_value_s, today_duration_s,
                        total_value_s, total_duration_s,
                        target_value_s, income_s
                    ])

    def _export_bait_data(self, wb: Workbook):
        """Export bait data to Excel"""
        ws = wb.create_sheet(title="饵料库存")
        ws.append(["名称", "已购买", "已使用", "剩余"])

        bait_file = os.path.join(self.data_dir, 'bait.json')
        if os.path.exists(bait_file):
            with open(bait_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for item in data:
                    bait = BaitConsumption(**item)
                    ws.append([bait.name, bait.total_bought, bait.total_used, bait.remaining])

    def _export_storage_data(self, wb: Workbook):
        """Export storage data to Excel"""
        ws = wb.create_sheet(title="存储时长")
        ws.append(["角色名称", "剩余时长(分钟)"])

        storage_file = os.path.join(self.data_dir, 'storage.json')
        if os.path.exists(storage_file):
            with open(storage_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for item in data:
                    storage = StorageCharacter(**item)
                    ws.append([storage.name, storage.remaining_minutes])

    def _export_credentials(self, wb: Workbook):
        """Export account names to Excel (passwords not exported for security)"""
        ws = wb.create_sheet(title="账号列表")
        ws.append(["账号名称"])

        cred_file = os.path.join(self.data_dir, 'credentials.json')
        if os.path.exists(cred_file):
            with open(cred_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for item in data:
                    cred = AccountCredential(**item)
                    ws.append([cred.account_name])
